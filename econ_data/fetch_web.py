"""Generalized web scraper for economic data not available via FRED or API.

Each scraper is defined in config.yaml under its group with source: web.
Scraper definitions live in SCRAPERS below — each maps a series_id to a
function that returns [(date, value), ...] from a web page.

To add a new web-scraped series:
  1. Write a _scrape_xxx() function that returns [(date, value), ...]
  2. Register it in SCRAPERS with its series_id
  3. Add the series to config.yaml with source: web
"""

import re
from datetime import date, datetime, timedelta
from pathlib import Path

import requests
from bs4 import BeautifulSoup

from econ_data.fetch import Observation

IMPORT_DIR = Path("import_files")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
}


def _get_soup(url: str) -> BeautifulSoup:
    """Fetch a URL and return a BeautifulSoup object."""
    resp = requests.get(url, headers=HEADERS, timeout=15)
    resp.raise_for_status()
    return BeautifulSoup(resp.text, "html.parser")


# ---------------------------------------------------------------------------
# Individual scrapers — each returns [(date, value), ...]
# ---------------------------------------------------------------------------

def _scrape_mba_purchase() -> list:
    """Scrape MBA Purchase Index from Trading Economics calendar table."""
    soup = _get_soup("https://tradingeconomics.com/united-states/mba-purchase-index")
    table = soup.find("table", id="calendar")
    if not table:
        raise ValueError("No calendar table found on Trading Economics MBA page")

    results = []
    for row in table.find_all("tr")[1:]:  # skip header
        cells = [td.get_text(strip=True) for td in row.find_all("td")]
        if len(cells) < 6:
            continue

        # Columns: date, time, name, reference, actual, previous, consensus, forecast
        release_date_str = cells[0].strip()
        ref = cells[3].strip()    # e.g. "Mar/13"
        actual = cells[4].strip()
        if not actual:
            continue  # future row, no data yet

        try:
            value = float(actual.replace(",", ""))
        except ValueError:
            continue

        # Parse reference week: "Mar/20" means the week ending on that date
        # Use the release year since reference doesn't include year
        try:
            release_date = datetime.strptime(release_date_str, "%Y-%m-%d").date()
            ref_match = re.match(r"(\w+)/(\d+)", ref)
            if ref_match:
                month_str, day_str = ref_match.groups()
                ref_date = datetime.strptime(f"{month_str} {day_str} {release_date.year}", "%b %d %Y").date()
                # Handle year boundary (Dec reference in Jan release)
                if ref_date > release_date + timedelta(days=60):
                    ref_date = ref_date.replace(year=ref_date.year - 1)
            else:
                ref_date = release_date
        except ValueError:
            continue

        results.append((ref_date, value))

    return results


def _scrape_wti_crude() -> list:
    """Fetch WTI Crude Oil futures (CL=F) daily close from Yahoo Finance API."""
    return _scrape_yahoo("CL=F", "5d")


def _scrape_dgs10() -> list:
    """Fetch 10-Year Treasury yield (^TNX) daily close from Yahoo Finance API.

    Forward-fills gaps (holidays, early closes) so every weekday has a value.
    Bond markets close on Good Friday, MLK Day, Presidents Day, etc. but
    mortgage rates are still quoted — we need a yield for spread calculations.
    """
    data = _scrape_yahoo("%5ETNX", "5d")
    return _forward_fill_weekdays(data)


def _forward_fill_weekdays(data: list) -> list:
    """Fill missing weekdays by carrying forward the prior value.

    Takes [(date, value), ...] sorted by date.  Inserts entries for any
    weekday gaps (holidays) between the first and last date, and also
    fills forward through yesterday if the last data point is older.
    """
    if not data:
        return data

    yesterday = date.today() - timedelta(days=1)
    by_date = dict(data)
    start = data[0][0]
    end = max(data[-1][0], yesterday)

    filled = []
    last_val = data[0][1]
    d = start
    while d <= end:
        if d.weekday() < 5:  # Mon-Fri
            if d in by_date:
                last_val = by_date[d]
            filled.append((d, last_val))
        d += timedelta(days=1)

    return filled


def _scrape_yahoo(symbol: str, range: str) -> list:
    """Fetch daily close from Yahoo Finance chart API."""
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
    resp = requests.get(url, params={"range": range, "interval": "1d"},
                        headers=HEADERS, timeout=15)
    resp.raise_for_status()
    data = resp.json()
    result = data["chart"]["result"][0]
    timestamps = result["timestamp"]
    closes = result["indicators"]["quote"][0]["close"]

    results = []
    for ts, close in zip(timestamps, closes):
        if close is None:
            continue
        obs_date = datetime.utcfromtimestamp(ts).date()
        results.append((obs_date, round(close, 2)))
    return results


# ---------------------------------------------------------------------------
# Import file parsers — for historical backfill from drop-folder spreadsheets
# ---------------------------------------------------------------------------

def _import_mba_purchase(last: date = None) -> list:
    """Import MBA Purchase Index history from grid-format spreadsheet.

    The file has weeks as rows and years as columns. Unpivots into a flat series.
    Only imports years before the current year — current-year dates from the grid
    are approximate and conflict with the scraper's accurate reference dates.
    """
    import openpyxl

    files = sorted(IMPORT_DIR.glob("MBA*Purchase*.xlsx"))
    if not files:
        return []

    current_year = date.today().year
    results = []
    for fpath in files:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws = wb["weekly YoY"]

        header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        years = [int(h) for h in header[1:] if h is not None]

        for row in ws.iter_rows(min_row=2, values_only=True):
            base_date = row[0]
            if base_date is None:
                continue
            if isinstance(base_date, datetime):
                base_date = base_date.date()

            for i, year in enumerate(years):
                if year >= current_year:
                    continue  # scraper handles current year with accurate dates
                value = row[i + 1]
                if value is None:
                    continue
                obs_date = date(year, base_date.month, base_date.day)
                if last and obs_date <= last:
                    continue
                results.append((obs_date, float(value)))

        wb.close()

    return results


def _import_wti_crude(last: date = None) -> list:
    """Backfill WTI Crude Oil futures from Yahoo Finance (2 years daily)."""
    return _scrape_yahoo("CL=F", "2y")


# Maps series_id to import function (for historical backfill from files)
IMPORTERS = {
    "MBA_PURCHASE": _import_mba_purchase,
    "WTI_CRUDE": _import_wti_crude,
}


# ---------------------------------------------------------------------------
# Scraper registry — maps series_id to (scrape_function, display_name)
# ---------------------------------------------------------------------------

SCRAPERS = {
    "MBA_PURCHASE": (_scrape_mba_purchase, "MBA Purchase Index"),
    "WTI_CRUDE": (_scrape_wti_crude, "WTI Crude Oil Futures"),
    "DGS10": (_scrape_dgs10, "10-Year Treasury Yield"),
}

# Cooldown: days to wait after last observation before scraping again
COOLDOWN = {
    "MBA_PURCHASE": 5,   # weekly, check a few days after last reading
    "WTI_CRUDE": 0,      # daily, always check
    "DGS10": 0,          # daily, always check
}

DEFAULT_COOLDOWN = 7


def fetch_web(last_dates: dict = None) -> dict:
    """
    Run all registered web scrapers and return new observations.

    Returns {"new": [Observation, ...], "counts": {series_id: int}}
    matching the same contract as other fetchers.
    """
    if last_dates is None:
        last_dates = {}

    all_new = []
    counts = {sid: 0 for sid in SCRAPERS}

    for series_id, (scrape_fn, name) in SCRAPERS.items():
        last = last_dates.get(series_id)

        # Try import files first (historical backfill)
        import_fn = IMPORTERS.get(series_id)
        if import_fn:
            try:
                imported = import_fn(last=last)
                for obs_date, value in imported:
                    all_new.append(Observation(
                        series_id=series_id,
                        name=name,
                        date=obs_date,
                        value=float(value),
                    ))
                    counts[series_id] = counts.get(series_id, 0) + 1
                    if last is None or obs_date > last:
                        last = obs_date
            except Exception as e:
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                print(f"[{ts}] SKIPPED {series_id} import — {e}")

        # Cooldown check (after import, so new imports update `last`)
        cooldown = COOLDOWN.get(series_id, DEFAULT_COOLDOWN)
        if last and (date.today() - last).days <= cooldown:
            continue

        try:
            data = scrape_fn()
        except Exception as e:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print(f"[{ts}] SKIPPED {series_id} — {e}")
            if counts[series_id] == 0:
                counts[series_id] = -1
            continue

        for obs_date, value in data:
            if last and obs_date <= last:
                continue

            all_new.append(Observation(
                series_id=series_id,
                name=name,
                date=obs_date,
                value=float(value),
            ))
            counts[series_id] += 1

    return {"new": all_new, "counts": counts}
