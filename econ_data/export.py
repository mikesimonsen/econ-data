import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from econ_data.db import connect
from econ_data.store import DB_PATH, get_export_log, get_last_dates

# Maps user-facing data type names to DB/export details
DATA_TYPES = {
    "values":     {"label": "Values",           "suffix": ""},
    "period_pct": {"label": "Period % Change",  "suffix": " period_pct"},
    "yoy_pct":    {"label": "YoY % Change",     "suffix": " yoy_pct"},
    "period_pp":  {"label": "Period pp Change",  "suffix": " period_pp"},
    "yoy_pp":     {"label": "YoY pp Change",     "suffix": " yoy_pp"},
    "all":        {"label": "All",              "suffix": ""},
}


def _query_raw(series_ids: list, db_path: Path = DB_PATH) -> dict:
    """Return {series_id: {"name": str, "rows": [(date, value), ...]}}."""
    con = connect()
    placeholders = ",".join(["%s"] * len(series_ids))
    cur = con.execute(
        f"SELECT series_id, name, date::text, value FROM observations "
        f"WHERE series_id IN ({placeholders}) ORDER BY series_id, date",
        series_ids,
    )
    result = {}
    for series_id, name, date, value in cur.fetchall():
        if series_id not in result:
            result[series_id] = {"name": name, "rows": []}
        result[series_id]["rows"].append((date, value))
    return result


def _query_calc(series_ids: list, calc_type: str, db_path: Path = DB_PATH) -> dict:
    """Return {series_id: {"name": str, "rows": [(date, value), ...]}} for a calc type."""
    con = connect()
    placeholders = ",".join(["%s"] * len(series_ids))
    # Get the canonical (oldest-row) name per series. Some series have multiple
    # names in observations (FRED renamed them); pick deterministically.
    names = {}
    for sid, name in con.execute(
        f"SELECT DISTINCT ON (series_id) series_id, name FROM observations "
        f"WHERE series_id IN ({placeholders}) ORDER BY series_id, date",
        series_ids,
    ).fetchall():
        names[sid] = name

    cur = con.execute(
        f"SELECT series_id, date::text, value FROM calculated "
        f"WHERE series_id IN ({placeholders}) AND calc_type = %s ORDER BY series_id, date",
        series_ids + [calc_type],
    )
    result = {}
    for series_id, date, value in cur.fetchall():
        if series_id not in result:
            result[series_id] = {"name": names.get(series_id, series_id), "rows": []}
        result[series_id]["rows"].append((date, value))
    return result


def _query_all(series_ids: list, db_path: Path = DB_PATH) -> dict:
    """Return {series_id: {"name": str, "rows": [(date, value, period_chg, yoy_chg), ...]}}."""
    raw = _query_raw(series_ids, db_path)
    period_pct = _query_calc(series_ids, "period_pct", db_path)
    yoy_pct = _query_calc(series_ids, "yoy_pct", db_path)
    period_pp = _query_calc(series_ids, "period_pp", db_path)
    yoy_pp = _query_calc(series_ids, "yoy_pp", db_path)

    result = {}
    for series_id, info in raw.items():
        # Use pp if available, otherwise pct
        period_map = {}
        yoy_map = {}
        if series_id in period_pp:
            period_map = {d: v for d, v in period_pp[series_id]["rows"]}
        elif series_id in period_pct:
            period_map = {d: v for d, v in period_pct[series_id]["rows"]}
        if series_id in yoy_pp:
            yoy_map = {d: v for d, v in yoy_pp[series_id]["rows"]}
        elif series_id in yoy_pct:
            yoy_map = {d: v for d, v in yoy_pct[series_id]["rows"]}

        combined = []
        for date_str, value in info["rows"]:
            combined.append((
                date_str,
                value,
                period_map.get(date_str, ""),
                yoy_map.get(date_str, ""),
            ))
        result[series_id] = {"name": info["name"], "rows": combined}
    return result


def available_series(db_path: Path = DB_PATH) -> list:
    """Return list of (series_id, name) stored in the database. Deduplicated by series_id."""
    return connect().execute(
        "SELECT DISTINCT ON (series_id) series_id, name FROM observations "
        "ORDER BY series_id, date"
    ).fetchall()


def available_groups(db_path: Path = DB_PATH) -> dict:
    """Return {group_id: {"name": str, "series": [(series_id, name), ...]}}."""
    con = connect()
    groups = {}
    for group_id, name in con.execute("SELECT group_id, name FROM groups ORDER BY group_id").fetchall():
        groups[group_id] = {"name": name, "series": []}
    for group_id, series_id in con.execute(
        "SELECT gm.group_id, o.series_id "
        "FROM group_members gm "
        "JOIN (SELECT DISTINCT series_id FROM observations) o ON gm.series_id = o.series_id "
        "ORDER BY gm.group_id, gm.series_id"
    ).fetchall():
        if group_id in groups:
            series_name = con.execute(
                "SELECT name FROM observations WHERE series_id = %s "
                "ORDER BY date LIMIT 1",
                (series_id,)
            ).fetchone()[0]
            groups[group_id]["series"].append((series_id, series_name))
    return groups


def series_for_group(group_id: str, db_path: Path = DB_PATH) -> list:
    """Return [series_id, ...] belonging to a group."""
    rows = connect().execute(
        "SELECT series_id FROM group_members WHERE group_id = %s ORDER BY series_id",
        (group_id,)
    ).fetchall()
    return [r[0] for r in rows]


def freshness(db_path: Path = DB_PATH) -> dict:
    """Return freshness info for standalone series and groups."""
    last_dates = get_last_dates(db_path)
    export_log = get_export_log(db_path)
    groups = available_groups(db_path)
    all_series = available_series(db_path)

    grouped_ids = {sid for gdata in groups.values() for sid, _ in gdata["series"]}

    standalone = []
    for sid, name in all_series:
        if sid not in grouped_ids:
            latest = last_dates.get(sid)
            prev = export_log.get(sid, {}).get("last_date")
            standalone.append((sid, name, latest.isoformat() if latest else None, prev))

    group_info = {}
    for gid, gdata in groups.items():
        series_fresh = []
        for sid, sname in gdata["series"]:
            latest = last_dates.get(sid)
            prev = export_log.get(f"{gid}:{sid}", {}).get("last_date")
            series_fresh.append((sid, sname, latest.isoformat() if latest else None, prev))

        group_latest = max((s[2] for s in series_fresh if s[2]), default=None)
        group_prev = export_log.get(gid, {}).get("last_date")

        group_info[gid] = {
            "name": gdata["name"],
            "latest": group_latest,
            "prev": group_prev,
            "series": series_fresh,
        }

    return {"standalone": standalone, "groups": group_info}


# ── Export writers ──────────────────────────────────────────────


def to_csv(series_ids: list, output_dir: Path, data_type: str = "values",
           db_path: Path = DB_PATH) -> list:
    """Write one CSV per series. Returns list of written file paths."""
    output_dir.mkdir(parents=True, exist_ok=True)
    written = []
    suffix = DATA_TYPES[data_type]["suffix"]

    if data_type == "all":
        data = _query_all(series_ids, db_path)
        for series_id, info in data.items():
            latest = info["rows"][-1][0]
            path = output_dir / f"{series_id} data thru {latest}.csv"
            with open(path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["date", series_id, "Period % Chg", "YoY % Chg"])
                writer.writerows(info["rows"])
            written.append(path)
    elif data_type == "values":
        data = _query_raw(series_ids, db_path)
        for series_id, info in data.items():
            latest = info["rows"][-1][0]
            path = output_dir / f"{series_id} data thru {latest}.csv"
            with open(path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["date", series_id])
                writer.writerows(info["rows"])
            written.append(path)
    else:
        data = _query_calc(series_ids, data_type, db_path)
        for series_id, info in data.items():
            latest = info["rows"][-1][0]
            path = output_dir / f"{series_id}{suffix} data thru {latest}.csv"
            with open(path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["date", f"{series_id} {DATA_TYPES[data_type]['label']}"])
                writer.writerows(info["rows"])
            written.append(path)

    return written


def to_excel(series_ids: list, output_path: Path, data_type: str = "values",
             db_path: Path = DB_PATH) -> Path:
    """Write one Excel workbook with one sheet per series."""
    suffix = DATA_TYPES[data_type]["suffix"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center")

    if data_type == "all":
        data = _query_all(series_ids, db_path)
        for series_id, info in data.items():
            ws = wb.create_sheet(title=series_id)
            headers = ["Date", series_id, "Period % Chg", "YoY % Chg"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            for row in info["rows"]:
                ws.append(list(row))
            for i in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 16
            ws.freeze_panes = "A2"
    elif data_type == "values":
        data = _query_raw(series_ids, db_path)
        for series_id, info in data.items():
            ws = wb.create_sheet(title=series_id)
            ws.append(["Date", series_id])
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            for date_str, value in info["rows"]:
                ws.append([date_str, value])
            ws.column_dimensions[get_column_letter(1)].width = 14
            ws.column_dimensions[get_column_letter(2)].width = 16
            ws.freeze_panes = "A2"
    else:
        data = _query_calc(series_ids, data_type, db_path)
        for series_id, info in data.items():
            ws = wb.create_sheet(title=series_id)
            col_name = f"{series_id} {DATA_TYPES[data_type]['label']}"
            ws.append(["Date", col_name])
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            for date_str, value in info["rows"]:
                ws.append([date_str, value])
            ws.column_dimensions[get_column_letter(1)].width = 14
            ws.column_dimensions[get_column_letter(2)].width = 20
            ws.freeze_panes = "A2"

    # ── Contents tab (always first) ─────────────────────────
    raw_data = _query_raw(series_ids, db_path)
    dt_label = DATA_TYPES[data_type]["label"]

    contents = wb.create_sheet(title="Contents", index=0)

    # Title row
    title_font = Font(bold=True, size=14)
    contents.append(["Table of Contents"])
    contents["A1"].font = title_font
    contents.append([])

    # Data type description
    contents.append(["Data on each tab:", dt_label])
    contents["A3"].font = Font(bold=True)
    contents.append([])

    # Column headers
    header_row = 5
    guide_headers = ["Tab", "Series Code", "Description", "First Date", "Last Date", "Observations"]
    contents.append(guide_headers)
    for cell in contents[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # One row per data tab
    for sid in series_ids:
        info = raw_data.get(sid)
        if info and info["rows"]:
            first = info["rows"][0][0]
            last = info["rows"][-1][0]
            count = len(info["rows"])
            name = info["name"]
        else:
            first = last = ""
            count = 0
            name = sid
        contents.append([sid, sid, name, first, last, count])

    contents.column_dimensions["A"].width = 22
    contents.column_dimensions["B"].width = 22
    contents.column_dimensions["C"].width = 50
    contents.column_dimensions["D"].width = 14
    contents.column_dimensions["E"].width = 14
    contents.column_dimensions["F"].width = 14
    contents.freeze_panes = f"A{header_row + 1}"

    # ── Filename with date stamp ────────────────────────────
    latest = max(
        (info["rows"][-1][0] for info in data.values() if info["rows"]),
        default="unknown",
    )
    stem = output_path.stem
    output_path = output_path.with_name(f"{stem}{suffix} data thru {latest}{output_path.suffix}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
