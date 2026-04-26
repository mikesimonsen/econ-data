"""
Compare multiple series side by side, aligned on common dates.
"""
import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from econ_data.db import connect
from econ_data.store import DB_PATH

DATA_TYPES = {
    "values":     {"label": "Values",          "table": "observations", "suffix": ""},
    "period_pct": {"label": "Period % Change", "table": "calculated",   "suffix": " period_pct"},
    "yoy_pct":    {"label": "YoY % Change",    "table": "calculated",   "suffix": " yoy_pct"},
    "period_pp":  {"label": "Period pp Change", "table": "calculated",  "suffix": " period_pp"},
    "yoy_pp":     {"label": "YoY pp Change",    "table": "calculated",  "suffix": " yoy_pp"},
}


def search_series(query: str, db_path: Path = DB_PATH) -> list:
    """Search for series by keyword in ID or name. Returns [(series_id, name), ...]."""
    q = f"%{query}%"
    return connect().execute(
        "SELECT DISTINCT series_id, name FROM observations "
        "WHERE series_id ILIKE %s OR name ILIKE %s "
        "ORDER BY series_id",
        (q, q),
    ).fetchall()


def _query_aligned(series_ids: list, data_type: str, db_path: Path = DB_PATH) -> dict:
    """
    Query series and align on common dates.
    Returns {"headers": [name, ...], "series_ids": [...], "dates": [...], "columns": [[val, ...], ...]}
    """
    con = connect()

    # Get names
    names = {}
    placeholders = ",".join(["%s"] * len(series_ids))
    for sid, name in con.execute(
        f"SELECT DISTINCT series_id, name FROM observations WHERE series_id IN ({placeholders})",
        series_ids,
    ).fetchall():
        names[sid] = name

    # Get data per series
    series_data = {}
    for sid in series_ids:
        if data_type == "values":
            rows = con.execute(
                "SELECT date::text, value FROM observations WHERE series_id = %s ORDER BY date",
                (sid,),
            ).fetchall()
        else:
            rows = con.execute(
                "SELECT date::text, value FROM calculated WHERE series_id = %s AND calc_type = %s ORDER BY date",
                (sid, data_type),
            ).fetchall()
        series_data[sid] = dict(rows)

    # Find common dates (inner join)
    if not series_data:
        return {"headers": [], "series_ids": [], "dates": [], "columns": []}

    date_sets = [set(d.keys()) for d in series_data.values()]
    common_dates = sorted(set.intersection(*date_sets))

    # Build aligned columns
    columns = []
    for sid in series_ids:
        columns.append([series_data[sid][d] for d in common_dates])

    return {
        "headers": [names.get(sid, sid) for sid in series_ids],
        "series_ids": series_ids,
        "dates": common_dates,
        "columns": columns,
    }


def to_excel(series_ids: list, data_type: str, output_path: Path,
             name: str = None, db_path: Path = DB_PATH) -> Path:
    """Write a comparison Excel workbook with all series as columns on one sheet."""
    aligned = _query_aligned(series_ids, data_type, db_path)

    if not aligned["dates"]:
        return None

    dt_label = DATA_TYPES[data_type]["label"]
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center")

    # ── Contents tab ────────────────────────────────────────
    contents = wb.active
    contents.title = "Contents"

    title_font = Font(bold=True, size=14)
    label = name or "Series Comparison"
    contents.append([label])
    contents["A1"].font = title_font
    contents.append([])
    contents.append(["Data on comparison tab:", dt_label])
    contents["A3"].font = Font(bold=True)
    contents.append([f"Common date range: {aligned['dates'][0]} to {aligned['dates'][-1]}"])
    contents.append([f"Observations: {len(aligned['dates'])}"])
    contents.append([])

    guide_headers = ["Column", "Series Code", "Description"]
    contents.append(guide_headers)
    for cell in contents[7]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for i, sid in enumerate(aligned["series_ids"]):
        contents.append([get_column_letter(i + 2), sid, aligned["headers"][i]])

    contents.column_dimensions["A"].width = 10
    contents.column_dimensions["B"].width = 22
    contents.column_dimensions["C"].width = 50

    # ── Data tab ────────────────────────────────────────────
    ws = wb.create_sheet(title="Comparison")

    # Headers: Date + series names
    row_headers = ["Date"] + aligned["headers"]
    ws.append(row_headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows
    for i, date in enumerate(aligned["dates"]):
        row = [date] + [col[i] for col in aligned["columns"]]
        ws.append(row)

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 14
    for j in range(len(aligned["headers"])):
        ws.column_dimensions[get_column_letter(j + 2)].width = max(
            16, len(aligned["headers"][j]) + 2
        )
    ws.freeze_panes = "B2"

    # ── Filename ────────────────────────────────────────────
    suffix = DATA_TYPES[data_type]["suffix"]
    latest = aligned["dates"][-1]
    stem = name or "comparison"
    output_path = output_path.parent / f"{stem}{suffix} data thru {latest}.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def to_csv(series_ids: list, data_type: str, output_path: Path,
           name: str = None, db_path: Path = DB_PATH) -> Path:
    """Write a comparison CSV with all series as columns."""
    aligned = _query_aligned(series_ids, data_type, db_path)

    if not aligned["dates"]:
        return None

    suffix = DATA_TYPES[data_type]["suffix"]
    latest = aligned["dates"][-1]
    stem = name or "comparison"
    output_path = output_path.parent / f"{stem}{suffix} data thru {latest}.csv"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["date"] + aligned["headers"])
        for i, date in enumerate(aligned["dates"]):
            writer.writerow([date] + [col[i] for col in aligned["columns"]])

    return output_path
